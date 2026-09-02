from pathlib import Path

from openpyxl import load_workbook

from src.final_test_report_generator import (
    generate_final_test_report,
)


def test_generates_formal_test_report(
    tmp_path: Path,
) -> None:
    """
    Verifies that the formal test report contains cover information,
    rule comparison data, and macros without approved rule coverage.
    """

    report_path = (
        tmp_path
        / "TSN2001_Compiler_Switches_Test_Report.xlsx"
    )

    test_report = {
        "test_id": "TSN2001",
        "title": "Compiler Switches Configuration",
        "executed_by": "Luis Fernando Vallejo Piña",
        "execution_date": "01-Sep-26",
        "mcu": "CYT2B73CAS",
        "expected_compiler": (
            "IAR C/C++ Compiler for ARM"
        ),
        "test_criteria": (
            "Verify compiler switches for Core1 Release."
        ),
        "project_identifier": "P09062 DMS",
    }

    makefile_data = {
        "Compiler": "iccarm.exe",
    }

    rule_results = [
        {
            "rule_id": "REL-001",
            "rule_type": "Macro",
            "macro": "DET_DEBUG_ENABLED",
            "description": (
                "DET debug support is enabled for Release."
            ),
            "expected_state": "Enabled",
            "expected_value": "",
            "expected_result": "",
            "actual_state": "Enabled",
            "actual_result": "",
            "resolved_value": "1u",
            "verdict": "PASS",
        },
        {
            "rule_id": "REL-033",
            "rule_type": "Expression",
            "macro": (
                "(DET_DEBUG_ENABLED == STD_ON) && "
                "(DET_DLTFILTERSIZE > 0)"
            ),
            "description": (
                "DET DLT filter code must remain inactive."
            ),
            "expected_state": "",
            "expected_value": "",
            "expected_result": "False",
            "actual_state": "False",
            "actual_result": "False",
            "occurrences": 3,
            "files_affected": 1,
            "resolved_value": "",
            "verdict": "PASS",
        },
    ]

    macro_resolutions = [
        {
            "primary_macro": "DET_DEBUG_ENABLED",
            "occurrences": 15,
            "files_affected": 2,
        },
    ]

    macro_verdict_coverage = [
        {
            "macro": "DET_DEBUG_ENABLED",
            "rule_verdict": "PASS",
            "coverage_status": "Rule configured",
            "example_source_file": (
                "C:/Project/Source/Core1/Det.c"
            ),
            "example_line": 69,
            "example_expression": (
                "DET_DEBUG_ENABLED == STD_ON"
            ),
            "actual_state": "Enabled",
        },
        {
            "macro": "VSECPRIM_AES128_ENABLED",
            "rule_verdict": "NOT_APPLICABLE",
            "coverage_status": (
                "No approved macro rule configured"
            ),
            "example_source_file": (
                "C:/Project/Source/Core1/ESLib.c"
            ),
            "example_line": 575,
            "example_expression": (
                "VSECPRIM_AES128_ENABLED == STD_ON"
            ),
            "actual_state": "Enabled",
        },
        {
            "macro": "UNKNOWN_FEATURE",
            "rule_verdict": "NOT_APPLICABLE",
            "coverage_status": (
                "No approved macro rule configured"
            ),
            "example_source_file": (
                "C:/Project/Source/Core1/Unknown.c"
            ),
            "example_line": 50,
            "example_expression": (
                "UNKNOWN_FEATURE == STD_ON"
            ),
            "actual_state": "Unresolved",
        },
    ]

    generated_report_path = generate_final_test_report(
        output_path=report_path,
        project_name="Release_DMS",
        core="Core1",
        build_mode="Release",
        project_root=Path("C:/Project"),
        test_report=test_report,
        makefile_data=makefile_data,
        rule_results=rule_results,
        macro_resolutions=macro_resolutions,
        macro_verdict_coverage=macro_verdict_coverage,
    )

    assert generated_report_path.exists()

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook["TSN2001 - Test Report"]

    assert worksheet["A1"].value == "TEST REPORT"

    assert (
        worksheet["A2"].value
        == "TSN2001 - Compiler Switches Configuration"
    )

    assert worksheet["A3"].value == "PROJECT INFORMATION"

    assert worksheet["A5"].value == "Test Executed By"

    assert worksheet["B5"].value == (
        "Luis Fernando Vallejo Piña"
    )

    assert worksheet["D5"].value == "Test Execution Date"
    assert worksheet["E5"].value == "01-Sep-26"

    assert worksheet["A6"].value == "Project"
    assert worksheet["B6"].value == "P09062 DMS"

    assert worksheet["D6"].value == "Verdict"
    assert worksheet["E6"].value == "PASS"

    assert (
        worksheet["A12"].value
        == "2. EVIDENCE COMPARISON TABLE (SUMMARY BY SWITCH)"
    )

    assert (
        worksheet["A13"].value
        == "Switch / Conditional Expression"
    )

    assert worksheet["A14"].value == "DET_DEBUG_ENABLED"
    assert worksheet["C14"].value == 15
    assert worksheet["D14"].value == 2
    assert worksheet["E14"].value == "PASS"

    assert (
        worksheet["A15"].value
        == (
            "(DET_DEBUG_ENABLED == STD_ON) && "
            "(DET_DLTFILTERSIZE > 0)"
        )
    )

    assert worksheet["C15"].value == 3
    assert worksheet["D15"].value == 1
    assert worksheet["E15"].value == "PASS"

    assert (
        worksheet["A18"].value
        == (
            "3. RELEVANT MACROS WITHOUT APPROVED RULE "
            "COVERAGE"
        )
    )

    assert worksheet["A19"].value == "File"
    assert worksheet["B19"].value == "Line and Expression"
    assert worksheet["C19"].value == "#if Result"

    assert worksheet["A20"].value == str(
        Path("Source") / "Core1" / "ESLib.c"
    )

    assert (
        worksheet["B20"].value
        == (
            "575: #if "
            "VSECPRIM_AES128_ENABLED == STD_ON"
        )
    )

    assert worksheet["C20"].value == "Enabled"

    assert worksheet["A21"].value == str(
        Path("Source") / "Core1" / "Unknown.c"
    )

    assert (
        worksheet["B21"].value
        == "50: #if UNKNOWN_FEATURE == STD_ON"
    )

    assert worksheet["C21"].value == "Unresolved"

    workbook.close()


def test_generates_fail_verdict_when_rule_fails(
    tmp_path: Path,
) -> None:
    """
    Verifies that one failed rule produces a FAIL verdict on the cover.
    """

    report_path = tmp_path / "failed_test_report.xlsx"

    generated_report_path = generate_final_test_report(
        output_path=report_path,
        project_name="Release_DMS",
        core="Core1",
        build_mode="Release",
        project_root=Path("."),
        test_report={
            "test_id": "TSN2001",
            "title": "Compiler Switches Configuration",
            "executed_by": "Tester",
            "execution_date": "01-Sep-26",
            "mcu": "MCU",
            "expected_compiler": "Compiler",
            "test_criteria": "Criterion",
            "project_identifier": "Project",
        },
        makefile_data={},
        rule_results=[
            {
                "rule_id": "REL-001",
                "rule_type": "Macro",
                "macro": "FEATURE_X",
                "description": "Feature X must be disabled.",
                "expected_state": "Disabled",
                "actual_state": "Enabled",
                "resolved_value": "1u",
                "verdict": "FAIL",
            }
        ],
        macro_resolutions=[],
        macro_verdict_coverage=[],
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook["TSN2001 - Test Report"]

    assert worksheet["E6"].value == "FAIL"

    workbook.close()