from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.makefile_parser import parse_makefile
from src.preprocessor_parser import find_preprocessor_directives
from src.report_generator import generate_excel_report
from src.switch_classifier import (
    classify_preprocessor_finding,
)


def get_project_root() -> Path:
    """
    Returns the root folder of the automation_tool project.
    """

    return Path(__file__).parent.parent


def get_summary_value(
    worksheet: Any,
    label: str,
) -> Any:
    """
    Returns the value in column B for the row whose column A matches
    the requested Summary label.
    """

    for row_number in range(
        1,
        worksheet.max_row + 1,
    ):
        if (
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
            == label
        ):
            return worksheet.cell(
                row=row_number,
                column=2,
            ).value

    raise AssertionError(
        f"Summary label not found: {label}"
    )


def test_generate_excel_report_with_controlled_data(
    tmp_path: Path,
) -> None:
    """
    Verifies that an Excel report is generated with the expected
    worksheets and controlled data.
    """

    project_root = get_project_root()

    fixture_makefile = (
        project_root
        / "tests"
        / "fixtures"
        / "compile_opt_example.mk"
    )

    fixture_c_file = (
        project_root
        / "tests"
        / "fixtures"
        / "ExampleModule.c"
    )

    makefile_data = parse_makefile(
        fixture_makefile,
        build_mode="Release",
    )

    findings = find_preprocessor_directives(
        fixture_c_file
    )

    report_path = (
        tmp_path
        / "compiler_switches_report.xlsx"
    )

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Controlled_Project",
        core="Core1",
        build_mode="Release",
        project_root=project_root,
        makefile_data=makefile_data,
        source_file_count=3,
        findings=findings,
    )

    assert generated_report_path.exists()

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    assert workbook.sheetnames == [
        "Summary",
        "Compiler Switches",
        "Preprocessor Conditions",
        "Filtered Compiler Switches",
        "Excluded Conditions",
        "Relevant Switch Summary",
        "Macro Resolution Summary",
        "Expression Evaluation",
        "Unresolved Expr Summary",
        "Rule Verdicts",
        "Macro Verdict Coverage",
    ]

    summary_sheet = workbook["Summary"]

    switches_sheet = workbook[
        "Compiler Switches"
    ]

    conditions_sheet = workbook[
        "Preprocessor Conditions"
    ]

    filtered_switches_sheet = workbook[
        "Filtered Compiler Switches"
    ]

    excluded_conditions_sheet = workbook[
        "Excluded Conditions"
    ]

    relevant_switch_summary_sheet = workbook[
        "Relevant Switch Summary"
    ]

    macro_resolution_sheet = workbook[
        "Macro Resolution Summary"
    ]

    expression_evaluation_sheet = workbook[
        "Expression Evaluation"
    ]

    unresolved_expression_sheet = workbook[
        "Unresolved Expr Summary"
    ]

    rule_verdicts_sheet = workbook[
        "Rule Verdicts"
    ]

    macro_verdict_coverage_sheet = workbook[
        "Macro Verdict Coverage"
    ]

    assert summary_sheet["A1"].value == "Metric"
    assert summary_sheet["B1"].value == "Value"

    assert switches_sheet["A1"].value == "Project"
    assert switches_sheet["J1"].value == "Entry Type"
    assert switches_sheet.max_row == 6

    assert conditions_sheet["A1"].value == "Source File"
    assert conditions_sheet["E1"].value == "Expression"
    assert conditions_sheet.max_row == 6

    assert (
        filtered_switches_sheet["A1"].value
        == "Source File"
    )

    assert (
        filtered_switches_sheet["G1"].value
        == "Category"
    )

    assert (
        excluded_conditions_sheet["A1"].value
        == "Source File"
    )

    assert (
        excluded_conditions_sheet["H1"].value
        == "Filter Reason"
    )

    assert (
        relevant_switch_summary_sheet["A1"].value
        == "Category"
    )

    assert (
        relevant_switch_summary_sheet["B1"].value
        == "Primary Macro"
    )

    assert (
        relevant_switch_summary_sheet["C1"].value
        == "Occurrences"
    )

    assert (
        relevant_switch_summary_sheet["D1"].value
        == "Files Affected"
    )

    assert (
        macro_resolution_sheet["A1"].value
        == "Category"
    )

    assert (
        macro_resolution_sheet["B1"].value
        == "Primary Macro"
    )

    assert (
        macro_resolution_sheet["E1"].value
        == "Resolved Value"
    )

    assert (
        macro_resolution_sheet["H1"].value
        == "Resolution Chain"
    )

    assert (
        macro_resolution_sheet["I1"].value
        == "Primary Definition Source"
    )

    assert (
        macro_resolution_sheet["M1"].value
        == "Terminal Definition Source"
    )

    assert (
        macro_resolution_sheet["T1"].value
        == "Primary Definition Condition"
    )

    assert (
        macro_resolution_sheet["U1"].value
        == "Resolved Primary Definition Condition"
    )

    assert (
        macro_resolution_sheet["V1"].value
        == "Primary Definition Condition Evaluation"
    )

    assert (
        macro_resolution_sheet["W1"].value
        == "Primary Definition Selection Reason"
    )

    assert (
        expression_evaluation_sheet["A1"].value
        == "Source File"
    )

    assert (
        expression_evaluation_sheet["F1"].value
        == "Original Expression"
    )

    assert (
        expression_evaluation_sheet["G1"].value
        == "Resolved Expression"
    )

    assert (
        expression_evaluation_sheet["I1"].value
        == "Verdict"
    )

    assert (
        expression_evaluation_sheet["L1"].value
        == "Error Message"
    )

    assert (
        unresolved_expression_sheet["A1"].value
        == "Error Type"
    )

    assert (
        unresolved_expression_sheet["B1"].value
        == "Issue Key"
    )

    assert (
        unresolved_expression_sheet["D1"].value
        == "Occurrences"
    )

    assert (
        unresolved_expression_sheet["F1"].value
        == "Error Message"
    )

    assert (
        rule_verdicts_sheet["A1"].value
        == "Rule ID"
    )

    assert (
        rule_verdicts_sheet["B1"].value
        == "Rule Type"
    )

    assert (
        rule_verdicts_sheet["C1"].value
        == "Macro / Expression"
    )

    assert (
        rule_verdicts_sheet["E1"].value
        == "Expected State"
    )

    assert (
        rule_verdicts_sheet["F1"].value
        == "Expected Value"
    )

    assert (
        rule_verdicts_sheet["G1"].value
        == "Expected Result"
    )

    assert (
        rule_verdicts_sheet["N1"].value
        == "Verdict"
    )

    assert (
        rule_verdicts_sheet["O1"].value
        == "Reason"
    )

    assert (
        rule_verdicts_sheet["P1"].value
        == "Resolution Chain"
    )

    assert (
        rule_verdicts_sheet["Q1"].value
        == "Primary Definition Source"
    )

    assert (
        rule_verdicts_sheet["U1"].value
        == "Primary Definition Condition"
    )

    assert (
        rule_verdicts_sheet["X1"].value
        == "Primary Definition Selection Reason"
    )

    assert (
        macro_verdict_coverage_sheet["A1"].value
        == "Category"
    )

    assert (
        macro_verdict_coverage_sheet["B1"].value
        == "Macro"
    )

    assert (
        macro_verdict_coverage_sheet["K1"].value
        == "Rule Verdict"
    )

    assert (
        macro_verdict_coverage_sheet["L1"].value
        == "Coverage Status"
    )

    workbook.close()


def test_summary_includes_classification_metrics(
    tmp_path: Path,
) -> None:
    """
    Verifies that Summary contains overall, exclusion, category,
    expression-evaluation, and rule-verdict metrics.
    """

    findings = [
        {
            "path": "src/Example.c",
            "file_name": "Example.c",
            "line_number": 10,
            "directive": "#if",
            "expression": "FEATURE_X",
            "macros": ["FEATURE_X"],
        },
        {
            "path": "src/Example.c",
            "file_name": "Example.c",
            "line_number": 20,
            "directive": "#ifdef",
            "expression": "__GNUC__",
            "macros": ["__GNUC__"],
        },
        {
            "path": "src/Example.h",
            "file_name": "Example.h",
            "line_number": 1,
            "directive": "#ifndef",
            "expression": "EXAMPLE_H",
            "macros": ["EXAMPLE_H"],
        },
        {
            "path": "src/Example.c",
            "file_name": "Example.c",
            "line_number": 30,
            "directive": "#if",
            "expression": "VALUE > 0",
            "macros": ["VALUE"],
        },
    ]

    classified_findings = [
        finding | classify_preprocessor_finding(
            finding
        )
        for finding in findings
    ]

    expression_evaluations = [
        {
            "evaluation_status": "Evaluated",
            "evaluation": True,
        },
        {
            "evaluation_status": "Evaluated",
            "evaluation": False,
        },
        {
            "evaluation_status": "Unresolved",
            "evaluation": None,
        },
    ]

    rule_summary = {
        "total_rules": 4,
        "pass_count": 1,
        "fail_count": 1,
        "review_count": 1,
        "not_applicable_count": 1,
    }

    report_path = (
        tmp_path
        / "summary_metrics_report.xlsx"
    )

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Metrics_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=2,
        findings=classified_findings,
        expression_evaluations=expression_evaluations,
        rule_summary=rule_summary,
        macro_coverage_summary={
            "total_macros": 4,
            "covered_macros": 2,
            "uncovered_macros": 2,
            "coverage_percentage": 50.0,
        },
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    summary_sheet = workbook["Summary"]

    assert get_summary_value(
        summary_sheet,
        "Total preprocessor directives analyzed",
    ) == 4

    assert get_summary_value(
        summary_sheet,
        "Relevant compiler switches",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Excluded conditions",
    ) == 2

    assert get_summary_value(
        summary_sheet,
        "Pending OTHER conditions for review",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Relevant expressions evaluated",
    ) == 2

    assert get_summary_value(
        summary_sheet,
        "Active branches",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Inactive branches",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Unresolved expressions",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Rules evaluated",
    ) == 4

    assert get_summary_value(
        summary_sheet,
        "Rule verdicts - PASS",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Rule verdicts - FAIL",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Rule verdicts - REVIEW",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Rule verdicts - NOT_APPLICABLE",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Relevant macros covered by rules",
    ) == 2

    assert get_summary_value(
        summary_sheet,
        "Relevant macros without approved rules",
    ) == 2

    assert get_summary_value(
        summary_sheet,
        "Macro rule coverage percentage",
    ) == 50.0

    assert get_summary_value(
        summary_sheet,
        "Header guard",
    ) == 1

    assert get_summary_value(
        summary_sheet,
        "Toolchain or architecture condition",
    ) == 1

    workbook.close()


def test_relevant_switch_summary_groups_repeated_macros(
    tmp_path: Path,
) -> None:
    """
    Verifies that repeated relevant conditions are consolidated into
    one summary row per category and primary macro.
    """

    findings = [
        {
            "path": "src/BswM.c",
            "file_name": "BswM.c",
            "line_number": 10,
            "directive": "#if",
            "expression": (
                "BSWM_ENABLE_CANSM == STD_ON"
            ),
            "macros": [
                "BSWM_ENABLE_CANSM",
                "STD_ON",
            ],
            "category": "FEATURE",
            "is_relevant": True,
            "filter_reason": "",
        },
        {
            "path": "src/BswM.c",
            "file_name": "BswM.c",
            "line_number": 25,
            "directive": "#if",
            "expression": (
                "BSWM_ENABLE_CANSM == STD_ON"
            ),
            "macros": [
                "BSWM_ENABLE_CANSM",
                "STD_ON",
            ],
            "category": "FEATURE",
            "is_relevant": True,
            "filter_reason": "",
        },
        {
            "path": "src/BswM_Cfg.h",
            "file_name": "BswM_Cfg.h",
            "line_number": 40,
            "directive": "#ifdef",
            "expression": "BSWM_ENABLE_CANSM",
            "macros": [
                "BSWM_ENABLE_CANSM",
            ],
            "category": "FEATURE",
            "is_relevant": True,
            "filter_reason": "",
        },
        {
            "path": "src/Det.c",
            "file_name": "Det.c",
            "line_number": 90,
            "directive": "#if",
            "expression": (
                "DET_DEBUG_ENABLED == STD_ON"
            ),
            "macros": [
                "DET_DEBUG_ENABLED",
                "STD_ON",
            ],
            "category": "DEBUG",
            "is_relevant": True,
            "filter_reason": "",
        },
    ]

    report_path = (
        tmp_path
        / "relevant_switch_summary.xlsx"
    )

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Summary_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=2,
        findings=findings,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook[
        "Relevant Switch Summary"
    ]

    assert worksheet.max_row == 3

    assert worksheet["A2"].value == "FEATURE"
    assert worksheet["B2"].value == "BSWM_ENABLE_CANSM"
    assert worksheet["C2"].value == 3
    assert worksheet["D2"].value == 2

    assert worksheet["A3"].value == "DEBUG"
    assert worksheet["B3"].value == "DET_DEBUG_ENABLED"
    assert worksheet["C3"].value == 1
    assert worksheet["D3"].value == 1

    workbook.close()


def test_macro_resolution_summary_writes_resolution_evidence(
    tmp_path: Path,
) -> None:
    """
    Verifies that Macro Resolution Summary writes one row with macro
    resolution data and primary/terminal definition evidence.
    """

    report_path = (
        tmp_path
        / "macro_resolution_report.xlsx"
    )

    macro_resolutions = [
        {
            "category": "FEATURE",
            "primary_macro": "FEATURE_X",
            "occurrences": 3,
            "files_affected": 2,
            "resolved_value": "1",
            "effective_state": "Enabled",
            "resolution_status": "Resolved",
            "resolution_chain_text": (
                "FEATURE_X -> STD_ON -> 1"
            ),
            "primary_definition_source": "Project_Cfg.h",
            "primary_definition_line": 45,
            "primary_definition_source_type": (
                "generated_config"
            ),
            "primary_definition_priority": 4,
            "terminal_definition_source": "Std_Types.h",
            "terminal_definition_line": 20,
            "terminal_definition_source_type": "header",
            "terminal_definition_priority": 1,
            "example_expression": (
                "FEATURE_X == STD_ON"
            ),
            "example_source_file": "src/Example.c",
            "example_line": 10,
            "primary_definition_condition": (
                "FEATURE_SELECTOR == STD_OFF"
            ),
            "resolved_primary_definition_condition": (
                "0u == 0u"
            ),
            "primary_definition_condition_evaluation": True,
            "primary_definition_selection_reason": (
                "Conditional branch evaluated as active"
            ),
        }
    ]

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Resolution_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=1,
        findings=[],
        macro_resolutions=macro_resolutions,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook[
        "Macro Resolution Summary"
    ]

    assert worksheet.max_row == 2

    assert worksheet["A2"].value == "FEATURE"
    assert worksheet["B2"].value == "FEATURE_X"
    assert worksheet["C2"].value == 3
    assert worksheet["D2"].value == 2
    assert worksheet["E2"].value == "1"
    assert worksheet["F2"].value == "Enabled"

    assert (
        worksheet["H2"].value
        == "FEATURE_X -> STD_ON -> 1"
    )

    assert worksheet["I2"].value == "Project_Cfg.h"
    assert worksheet["J2"].value == 45
    assert worksheet["K2"].value == "generated_config"
    assert worksheet["L2"].value == 4

    assert worksheet["M2"].value == "Std_Types.h"
    assert worksheet["N2"].value == 20
    assert worksheet["O2"].value == "header"
    assert worksheet["P2"].value == 1

    assert (
        worksheet["Q2"].value
        == "FEATURE_X == STD_ON"
    )

    assert worksheet["R2"].value == "src/Example.c"
    assert worksheet["S2"].value == 10

    assert (
        worksheet["T2"].value
        == "FEATURE_SELECTOR == STD_OFF"
    )

    assert (
        worksheet["U2"].value
        == "0u == 0u"
    )

    assert worksheet["V2"].value == "True"

    assert (
        worksheet["W2"].value
        == "Conditional branch evaluated as active"
    )

    workbook.close()


def test_expression_evaluation_sheet_writes_results(
    tmp_path: Path,
) -> None:
    """
    Verifies that preprocessor expression evaluation evidence is
    written to the dedicated report worksheet.
    """

    report_path = (
        tmp_path
        / "expression_evaluation_report.xlsx"
    )

    expression_evaluations = [
        {
            "source_file": "src/Example.c",
            "file_name": "Example.c",
            "line_number": 42,
            "directive": "#if",
            "category": "FEATURE",
            "original_expression": (
                "FEATURE_X == STD_ON"
            ),
            "resolved_expression": "(1u == 1u)",
            "evaluation": True,
            "verdict": "Active branch",
            "evaluation_status": "Evaluated",
            "referenced_macros": [
                "FEATURE_X",
                "STD_ON",
            ],
            "error_message": "",
        }
    ]

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Evaluation_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=1,
        findings=[],
        expression_evaluations=expression_evaluations,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook[
        "Expression Evaluation"
    ]

    assert worksheet.max_row == 2
    assert worksheet["A2"].value == "src/Example.c"
    assert worksheet["B2"].value == "Example.c"
    assert worksheet["C2"].value == 42
    assert worksheet["D2"].value == "#if"
    assert worksheet["E2"].value == "FEATURE"

    assert (
        worksheet["F2"].value
        == "FEATURE_X == STD_ON"
    )

    assert worksheet["G2"].value == "(1u == 1u)"
    assert worksheet["H2"].value == "True"
    assert worksheet["I2"].value == "Active branch"
    assert worksheet["J2"].value == "Evaluated"

    assert (
        worksheet["K2"].value
        == "FEATURE_X, STD_ON"
    )

    assert worksheet["L2"].value in (None, "")

    workbook.close()


def test_unresolved_expression_summary_sheet_writes_results(
    tmp_path: Path,
) -> None:
    """
    Verifies that grouped unresolved-expression issues are written to
    the dedicated worksheet.
    """

    report_path = (
        tmp_path
        / "unresolved_expression_summary.xlsx"
    )

    unresolved_summary = [
        {
            "error_type": "Missing macro definition",
            "issue_key": "DEM_FEATURE_FAST",
            "category": "FEATURE",
            "occurrences": 4,
            "files_affected": 2,
            "error_message": (
                "Macro definition not found: "
                "DEM_FEATURE_FAST"
            ),
            "example_original_expression": (
                "DEM_FEATURE_FAST == STD_ON"
            ),
            "example_source_file": "src/Dem_A.c",
            "example_line": 1362,
            "example_directive": "#if",
        }
    ]

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Unresolved_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=1,
        findings=[],
        unresolved_expression_summary=unresolved_summary,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook[
        "Unresolved Expr Summary"
    ]

    assert worksheet.max_row == 2
    assert worksheet["A2"].value == "Missing macro definition"
    assert worksheet["B2"].value == "DEM_FEATURE_FAST"
    assert worksheet["C2"].value == "FEATURE"
    assert worksheet["D2"].value == 4
    assert worksheet["E2"].value == 2

    assert (
        worksheet["F2"].value
        == "Macro definition not found: "
        "DEM_FEATURE_FAST"
    )

    assert (
        worksheet["G2"].value
        == "DEM_FEATURE_FAST == STD_ON"
    )

    assert worksheet["H2"].value == "src/Dem_A.c"
    assert worksheet["I2"].value == 1362
    assert worksheet["J2"].value == "#if"

    workbook.close()


def test_rule_verdicts_sheet_writes_rule_results(
    tmp_path: Path,
) -> None:
    """
    Verifies that macro-rule results are written to the Rule Verdicts
    worksheet with the expected evidence columns.
    """

    report_path = (
        tmp_path
        / "rule_verdicts_report.xlsx"
    )

    rule_results = [
        {
            "rule_id": "RULE-001",
            "rule_type": "Macro",
            "macro": "FEATURE_X",
            "description": "Feature X must be enabled.",
            "expected_state": "Enabled",
            "expected_value": "",
            "expected_result": "",
            "actual_state": "Enabled",
            "actual_result": "",
            "occurrences": "",
            "files_affected": "",
            "resolved_value": "1u",
            "resolution_status": "Resolved",
            "verdict": "PASS",
            "reason": (
                "The actual macro state matches the expected state."
            ),
            "resolution_chain": (
                "FEATURE_X -> STD_ON -> 1u"
            ),
            "primary_definition_source": "Project_Cfg.h",
            "primary_definition_line": 45,
            "primary_definition_source_type": (
                "conditional_definition"
            ),
            "primary_definition_priority": 5,
            "primary_definition_condition": (
                "FEATURE_SELECTOR == STD_ON"
            ),
            "resolved_primary_definition_condition": (
                "1u == 1u"
            ),
            "primary_definition_condition_evaluation": True,
            "primary_definition_selection_reason": (
                "Conditional branch evaluated as active"
            ),
        },
        {
            "rule_id": "RULE-002",
            "rule_type": "Macro",
            "macro": "FEATURE_Y",
            "description": "Feature Y must be disabled.",
            "expected_state": "Disabled",
            "expected_value": "",
            "expected_result": "",
            "actual_state": "Enabled",
            "actual_result": "",
            "occurrences": "",
            "files_affected": "",
            "resolved_value": "1u",
            "resolution_status": "Resolved",
            "verdict": "FAIL",
            "reason": (
                "The actual macro state does not match the expected "
                "state."
            ),
        },
    ]

    rule_summary = {
        "total_rules": 2,
        "pass_count": 1,
        "fail_count": 1,
        "review_count": 0,
        "not_applicable_count": 0,
    }

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Rules_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=1,
        findings=[],
        rule_results=rule_results,
        rule_summary=rule_summary,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook["Rule Verdicts"]

    assert worksheet.max_row == 3

    # First macro rule.
    assert worksheet["A2"].value == "RULE-001"
    assert worksheet["B2"].value == "Macro"
    assert worksheet["C2"].value == "FEATURE_X"
    assert worksheet["D2"].value == "Feature X must be enabled."

    assert worksheet["E2"].value == "Enabled"
    assert worksheet["F2"].value in (None, "")
    assert worksheet["G2"].value in (None, "")

    assert worksheet["H2"].value == "Enabled"
    assert worksheet["I2"].value in (None, "")
    assert worksheet["J2"].value in (None, "")
    assert worksheet["K2"].value in (None, "")

    assert worksheet["L2"].value == "1u"
    assert worksheet["M2"].value == "Resolved"
    assert worksheet["N2"].value == "PASS"

    assert (
        worksheet["O2"].value
        == "The actual macro state matches the expected state."
    )

    assert (
        worksheet["P2"].value
        == "FEATURE_X -> STD_ON -> 1u"
    )

    assert worksheet["Q2"].value == "Project_Cfg.h"
    assert worksheet["R2"].value == 45

    assert (
        worksheet["S2"].value
        == "conditional_definition"
    )

    assert worksheet["T2"].value == 5

    assert (
        worksheet["U2"].value
        == "FEATURE_SELECTOR == STD_ON"
    )

    assert (
        worksheet["V2"].value
        == "1u == 1u"
    )

    assert worksheet["W2"].value == "True"

    assert (
        worksheet["X2"].value
        == "Conditional branch evaluated as active"
    )

    # Second macro rule.
    assert worksheet["A3"].value == "RULE-002"
    assert worksheet["B3"].value == "Macro"
    assert worksheet["C3"].value == "FEATURE_Y"
    assert worksheet["E3"].value == "Disabled"
    assert worksheet["H3"].value == "Enabled"
    assert worksheet["N3"].value == "FAIL"

    workbook.close()

def test_macro_verdict_coverage_sheet_writes_coverage_rows(
    tmp_path: Path,
) -> None:
    """
    Verifies that Macro Verdict Coverage writes rule coverage for
    configured and non-configured macros.
    """

    report_path = (
        tmp_path
        / "macro_verdict_coverage_report.xlsx"
    )

    coverage_rows = [
        {
            "category": "DEBUG",
            "macro": "DET_DEBUG_ENABLED",
            "occurrences": 15,
            "files_affected": 2,
            "actual_state": "Enabled",
            "resolved_value": "1u",
            "resolution_status": "Resolved",
            "rule_id": "REL-001",
            "expected_state": "Enabled",
            "expected_value": "",
            "rule_verdict": "PASS",
            "coverage_status": "Rule configured",
            "reason": (
                "The actual macro state matches the expected state."
            ),
            "resolution_chain": (
                "DET_DEBUG_ENABLED -> STD_ON -> 1u"
            ),
            "primary_definition_source": "Det_Cfg.h",
            "primary_definition_line": 62,
            "primary_definition_source_type": (
                "conditional_definition"
            ),
        },
        {
            "category": "FEATURE",
            "macro": "VSECPRIM_AES128_ENABLED",
            "occurrences": 4,
            "files_affected": 2,
            "actual_state": "Enabled",
            "resolved_value": "1u",
            "resolution_status": "Resolved",
            "rule_id": "",
            "expected_state": "",
            "expected_value": "",
            "rule_verdict": "NOT_APPLICABLE",
            "coverage_status": (
                "No approved macro rule configured"
            ),
            "reason": (
                "No approved macro rule is configured for this "
                "relevant macro."
            ),
            "resolution_chain": (
                "VSECPRIM_AES128_ENABLED -> STD_ON -> 1u"
            ),
            "primary_definition_source": "ESLib_Cfg.h",
            "primary_definition_line": 120,
            "primary_definition_source_type": (
                "project_source"
            ),
        },
    ]

    generated_report_path = generate_excel_report(
        output_path=report_path,
        project_name="Coverage_Project",
        core="Core1",
        build_mode="Release",
        project_root=get_project_root(),
        makefile_data={},
        source_file_count=1,
        findings=[],
        macro_verdict_coverage=coverage_rows,
    )

    workbook = load_workbook(
        generated_report_path,
        read_only=True,
    )

    worksheet = workbook[
        "Macro Verdict Coverage"
    ]

    assert worksheet.max_row == 3

    assert worksheet["A1"].value == "Category"
    assert worksheet["B1"].value == "Macro"
    assert worksheet["H1"].value == "Rule ID"
    assert worksheet["K1"].value == "Rule Verdict"
    assert worksheet["L1"].value == "Coverage Status"

    assert worksheet["A2"].value == "DEBUG"
    assert worksheet["B2"].value == "DET_DEBUG_ENABLED"
    assert worksheet["H2"].value == "REL-001"
    assert worksheet["I2"].value == "Enabled"
    assert worksheet["K2"].value == "PASS"
    assert worksheet["L2"].value == "Rule configured"

    assert worksheet["A3"].value == "FEATURE"
    assert worksheet["B3"].value == "VSECPRIM_AES128_ENABLED"
    assert worksheet["H3"].value in (None, "")
    assert worksheet["K3"].value == "NOT_APPLICABLE"

    assert (
        worksheet["L3"].value
        == "No approved macro rule configured"
    )

    workbook.close()