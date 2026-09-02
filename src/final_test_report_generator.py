from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.worksheet import Worksheet


DARK_BLUE_FILL = PatternFill(
    fill_type="solid",
    fgColor="203864",
)

SECTION_BLUE_FILL = PatternFill(
    fill_type="solid",
    fgColor="2F5597",
)

LABEL_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9E2F3",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

PASS_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

FAIL_FILL = PatternFill(
    fill_type="solid",
    fgColor="FCE4D6",
)

REVIEW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

TITLE_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=18,
)

SUBTITLE_FONT = Font(
    bold=True,
    color="FFFFFF",
    italic=True,
    size=11,
)

SECTION_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=12,
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

LABEL_FONT = Font(
    bold=True,
    color="1F1F1F",
)

THIN_SIDE = Side(
    style="thin",
    color="A6A6A6",
)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def generate_final_test_report(
    output_path: str | Path,
    project_name: str,
    core: str,
    build_mode: str,
    project_root: Path,
    test_report: dict[str, Any],
    makefile_data: dict[str, Any],
    rule_results: Iterable[dict[str, Any]],
    macro_resolutions: Iterable[dict[str, Any]],
    macro_verdict_coverage: Iterable[dict[str, Any]],
) -> Path:
    """
    Generates a compact formal test report.

    The report contains:
    1. Cover page / project information.
    2. Evidence comparison table summarized by approved rules.
    3. Relevant macros without approved rule coverage.

    Detailed technical evidence remains available separately in:
    compiler_switches_report.xlsx.
    """

    report_path = Path(output_path)

    report_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    rule_results_list = list(rule_results)

    macro_resolutions_list = list(
        macro_resolutions
    )

    macro_verdict_coverage_list = list(
        macro_verdict_coverage
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = _build_worksheet_title(
        test_report=test_report
    )

    _configure_worksheet(
        worksheet=worksheet,
    )

    _write_cover_section(
        worksheet=worksheet,
        project_name=project_name,
        core=core,
        build_mode=build_mode,
        test_report=test_report,
        makefile_data=makefile_data,
        rule_results=rule_results_list,
    )

    comparison_end_row = _write_comparison_table(
        worksheet=worksheet,
        start_row=12,
        rule_results=rule_results_list,
        macro_resolutions=macro_resolutions_list,
    )

    _write_uncovered_macro_table(
        worksheet=worksheet,
        start_row=comparison_end_row + 3,
        project_root=project_root,
        macro_verdict_coverage=(
            macro_verdict_coverage_list
        ),
    )

    workbook.save(report_path)
    workbook.close()

    return report_path


def _build_worksheet_title(
    test_report: dict[str, Any],
) -> str:
    """
    Builds an Excel-compatible worksheet title.

    Excel worksheet titles are limited to 31 characters.
    """

    test_id = str(
        test_report.get(
            "test_id",
            "TSN2001",
        )
    ).strip()

    worksheet_title = f"{test_id} - Test Report"

    return worksheet_title[:31]


def _configure_worksheet(
    worksheet: Worksheet,
) -> None:
    """
    Applies document-level worksheet settings.
    """

    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A13"

    worksheet.column_dimensions["A"].width = 80
    worksheet.column_dimensions["B"].width = 105
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["F"].width = 18


def _write_cover_section(
    worksheet: Worksheet,
    project_name: str,
    core: str,
    build_mode: str,
    test_report: dict[str, Any],
    makefile_data: dict[str, Any],
    rule_results: list[dict[str, Any]],
) -> None:
    """
    Writes the formal report cover and project information section.
    """

    test_id = str(
        test_report.get(
            "test_id",
            "TSN2001",
        )
    ).strip()

    report_title = str(
        test_report.get(
            "title",
            "Compiler Switches Configuration",
        )
    ).strip()

    executed_by = str(
        test_report.get(
            "executed_by",
            "Not configured",
        )
    ).strip()

    project_identifier = str(
        test_report.get(
            "project_identifier",
            project_name,
        )
    ).strip()

    mcu = str(
        test_report.get(
            "mcu",
            "Not configured",
        )
    ).strip()

    expected_compiler = str(
        test_report.get(
            "expected_compiler",
            "",
        )
    ).strip()

    test_criteria = str(
        test_report.get(
            "test_criteria",
            "Not configured",
        )
    ).strip()

    compiler = _get_report_compiler(
        makefile_data=makefile_data,
        expected_compiler=expected_compiler,
    )

    execution_date = _resolve_execution_date(
        test_report=test_report,
    )

    overall_verdict = _calculate_overall_verdict(
        rule_results=rule_results,
    )

    worksheet.merge_cells("A1:F1")
    worksheet["A1"] = "TEST REPORT"

    worksheet["A1"].fill = DARK_BLUE_FILL
    worksheet["A1"].font = TITLE_FONT
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells("A2:F2")
    worksheet["A2"] = f"{test_id} - {report_title}"

    worksheet["A2"].fill = DARK_BLUE_FILL
    worksheet["A2"].font = SUBTITLE_FONT
    worksheet["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[2].height = 20

    _write_section_title(
        worksheet=worksheet,
        row_number=3,
        title="PROJECT INFORMATION",
    )

    _write_label(
        worksheet=worksheet,
        cell_reference="A5",
        value="Test Executed By",
    )

    worksheet.merge_cells("B5:C5")
    worksheet["B5"] = executed_by

    _write_label(
        worksheet=worksheet,
        cell_reference="D5",
        value="Test Execution Date",
    )

    worksheet.merge_cells("E5:F5")
    worksheet["E5"] = execution_date

    _write_label(
        worksheet=worksheet,
        cell_reference="A6",
        value="Project",
    )

    worksheet.merge_cells("B6:C6")
    worksheet["B6"] = project_identifier

    _write_label(
        worksheet=worksheet,
        cell_reference="D6",
        value="Verdict",
    )

    worksheet.merge_cells("E6:F6")
    worksheet["E6"] = overall_verdict

    _apply_verdict_style(
        cell=worksheet["E6"],
        verdict=overall_verdict,
    )

    _write_label(
        worksheet=worksheet,
        cell_reference="A7",
        value="Configuration",
    )

    worksheet.merge_cells("B7:C7")
    worksheet["B7"] = build_mode

    _write_label(
        worksheet=worksheet,
        cell_reference="D7",
        value="Core",
    )

    worksheet.merge_cells("E7:F7")
    worksheet["E7"] = core

    _write_label(
        worksheet=worksheet,
        cell_reference="A8",
        value="MCU",
    )

    worksheet.merge_cells("B8:F8")
    worksheet["B8"] = mcu

    _write_label(
        worksheet=worksheet,
        cell_reference="A9",
        value="Compiler",
    )

    worksheet.merge_cells("B9:F9")
    worksheet["B9"] = compiler

    _write_label(
        worksheet=worksheet,
        cell_reference="A10",
        value="Test Criteria",
    )

    worksheet.merge_cells("B10:F10")
    worksheet["B10"] = test_criteria

    for row_number in range(5, 11):
        for column_number in range(1, 7):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

    worksheet.row_dimensions[10].height = 42


def _write_comparison_table(
    worksheet: Worksheet,
    start_row: int,
    rule_results: list[dict[str, Any]],
    macro_resolutions: list[dict[str, Any]],
) -> int:
    """
    Writes section 2: evidence comparison table summarized by switch.
    """

    _write_section_title(
        worksheet=worksheet,
        row_number=start_row,
        title="2. EVIDENCE COMPARISON TABLE (SUMMARY BY SWITCH)",
    )

    header_row = start_row + 1

    headers = [
        "Switch / Conditional Expression",
        "Justification / Comment",
        "Occurrences",
        "Files",
        "Result",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    macro_resolutions_by_name = {
        str(
            resolution.get(
                "primary_macro",
                "",
            )
        ): resolution
        for resolution in macro_resolutions
    }

    current_row = header_row + 1

    for rule_result in rule_results:
        rule_type = str(
            rule_result.get(
                "rule_type",
                "Macro",
            )
        )

        macro_or_expression = str(
            rule_result.get("macro", "")
        )

        occurrences, files_affected = (
            _get_rule_occurrence_data(
                rule_result=rule_result,
                rule_type=rule_type,
                macro_resolutions_by_name=(
                    macro_resolutions_by_name
                ),
            )
        )

        worksheet.cell(
            row=current_row,
            column=1,
            value=macro_or_expression,
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=_build_rule_comment(
                rule_result=rule_result,
            ),
        )

        worksheet.cell(
            row=current_row,
            column=3,
            value=occurrences,
        )

        worksheet.cell(
            row=current_row,
            column=4,
            value=files_affected,
        )

        verdict_cell = worksheet.cell(
            row=current_row,
            column=5,
            value=str(
                rule_result.get(
                    "verdict",
                    "",
                )
            ),
        )

        _apply_verdict_style(
            cell=verdict_cell,
            verdict=str(
                rule_result.get(
                    "verdict",
                    "",
                )
            ),
        )

        for column_number in range(1, 6):
            cell = worksheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        worksheet.row_dimensions[current_row].height = 50

        current_row += 1

    worksheet.auto_filter.ref = (
        f"A{header_row}:E{current_row - 1}"
    )

    return current_row - 1


def _write_uncovered_macro_table(
    worksheet: Worksheet,
    start_row: int,
    project_root: Path,
    macro_verdict_coverage: list[dict[str, Any]],
) -> None:
    """
    Writes section 3: relevant macros without approved rule coverage.

    This section intentionally shows only macros with:
    - Rule Verdict: NOT_APPLICABLE
    - Coverage Status: No approved macro rule configured

    These macros are visible and traceable but do not contribute to
    the final PASS/FAIL verdict because no approved expected behavior
    has been configured for them.
    """

    _write_section_title(
        worksheet=worksheet,
        row_number=start_row,
        title=(
            "3. RELEVANT MACROS WITHOUT APPROVED RULE COVERAGE"
        ),
    )

    header_row = start_row + 1

    headers = [
        "File",
        "Line and Expression",
        "#if Result",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=header_row,
            column=column_number,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    uncovered_rows = [
        coverage_row
        for coverage_row in macro_verdict_coverage
        if (
            coverage_row.get("rule_verdict")
            == "NOT_APPLICABLE"
        )
        and (
            coverage_row.get("coverage_status")
            == "No approved macro rule configured"
        )
    ]

    current_row = header_row + 1

    for coverage_row in uncovered_rows:
        source_file = _format_source_file(
            source_file=str(
                coverage_row.get(
                    "example_source_file",
                    "",
                )
            ),
            project_root=project_root,
        )

        line_number = coverage_row.get(
            "example_line",
            "",
        )

        expression = str(
            coverage_row.get(
                "example_expression",
                "",
            )
        )

        macro_name = str(
            coverage_row.get("macro", "")
        )

        actual_state = str(
            coverage_row.get(
                "actual_state",
                "",
            )
        )

        if expression:
            line_and_expression = (
                f"{line_number}: #if {expression}"
            ).strip()
        else:
            line_and_expression = (
                f"{line_number}: {macro_name}"
            ).strip()

        worksheet.cell(
            row=current_row,
            column=1,
            value=source_file,
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value=line_and_expression,
        )

        worksheet.cell(
            row=current_row,
            column=3,
            value=actual_state,
        )

        for column_number in range(1, 4):
            cell = worksheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        current_row += 1

    worksheet.auto_filter.ref = (
        f"A{header_row}:C{current_row - 1}"
    )

    worksheet.column_dimensions["A"].width = 80
    worksheet.column_dimensions["B"].width = 105
    worksheet.column_dimensions["C"].width = 18


def _write_section_title(
    worksheet: Worksheet,
    row_number: int,
    title: str,
) -> None:
    """
    Writes a merged dark-blue section title.
    """

    worksheet.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=6,
    )

    cell = worksheet.cell(
        row=row_number,
        column=1,
        value=title,
    )

    cell.fill = SECTION_BLUE_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    cell.border = THIN_BORDER
    worksheet.row_dimensions[row_number].height = 22


def _write_label(
    worksheet: Worksheet,
    cell_reference: str,
    value: str,
) -> None:
    """
    Writes one project-information label.
    """

    cell = worksheet[cell_reference]

    cell.value = value
    cell.fill = LABEL_FILL
    cell.font = LABEL_FONT
    cell.alignment = Alignment(
        vertical="center",
        wrap_text=True,
    )


def _get_report_compiler(
    makefile_data: dict[str, Any],
    expected_compiler: str,
) -> str:
    """
    Selects a readable compiler name for the formal report.
    """

    detected_compiler = str(
        makefile_data.get(
            "Compiler",
            "",
        )
    ).strip()

    if (
        detected_compiler
        and detected_compiler != "Not detected"
    ):
        return detected_compiler

    if expected_compiler:
        return expected_compiler

    return "Not detected"


def _resolve_execution_date(
    test_report: dict[str, Any],
) -> str:
    """
    Resolves report execution date from YAML metadata.

    The value 'auto' uses the local execution date.
    """

    configured_date = str(
        test_report.get(
            "execution_date",
            "auto",
        )
    ).strip()

    if configured_date.lower() == "auto":
        return date.today().strftime(
            "%d-%b-%y"
        )

    return configured_date


def _calculate_overall_verdict(
    rule_results: list[dict[str, Any]],
) -> str:
    """
    Calculates final verdict from configured rule results.

    Priority:
    FAIL > REVIEW > PASS.
    """

    verdicts = {
        str(result.get("verdict", ""))
        for result in rule_results
    }

    if "FAIL" in verdicts:
        return "FAIL"

    if "REVIEW" in verdicts:
        return "REVIEW"

    if "PASS" in verdicts:
        return "PASS"

    return "REVIEW"


def _get_rule_occurrence_data(
    rule_result: dict[str, Any],
    rule_type: str,
    macro_resolutions_by_name: dict[
        str,
        dict[str, Any],
    ],
) -> tuple[int | str, int | str]:
    """
    Returns occurrences and affected-file count for one rule.
    """

    if rule_type == "Expression":
        return (
            rule_result.get("occurrences", ""),
            rule_result.get("files_affected", ""),
        )

    macro_name = str(
        rule_result.get("macro", "")
    )

    macro_resolution = macro_resolutions_by_name.get(
        macro_name
    )

    if macro_resolution is None:
        return "", ""

    return (
        macro_resolution.get("occurrences", ""),
        macro_resolution.get("files_affected", ""),
    )


def _build_rule_comment(
    rule_result: dict[str, Any],
) -> str:
    """
    Builds a concise technical justification for the comparison table.
    """

    description = str(
        rule_result.get("description", "")
    ).strip()

    rule_type = str(
        rule_result.get(
            "rule_type",
            "Macro",
        )
    )

    if rule_type == "Expression":
        expected_result = str(
            rule_result.get(
                "expected_result",
                "",
            )
        )

        actual_result = str(
            rule_result.get(
                "actual_result",
                "",
            )
        )

        return (
            f"{description}\n"
            f"Expected result: {expected_result}. "
            f"Actual result: {actual_result}."
        )

    expected_state = str(
        rule_result.get(
            "expected_state",
            "",
        )
    ).strip()

    expected_value = str(
        rule_result.get(
            "expected_value",
            "",
        )
    ).strip()

    actual_state = str(
        rule_result.get(
            "actual_state",
            "",
        )
    ).strip()

    resolved_value = str(
        rule_result.get(
            "resolved_value",
            "",
        )
    ).strip()

    if expected_value:
        return (
            f"{description}\n"
            f"Expected value: {expected_value}. "
            f"Actual value: {resolved_value}."
        )

    if expected_state:
        return (
            f"{description}\n"
            f"Expected state: {expected_state}. "
            f"Actual state: {actual_state}."
        )

    return description


def _format_source_file(
    source_file: str,
    project_root: Path,
) -> str:
    """
    Converts an absolute source-file path into a project-relative path
    when possible.
    """

    if not source_file:
        return ""

    source_path = Path(source_file)

    try:
        return str(
            source_path.relative_to(project_root)
        )
    except ValueError:
        return source_file


def _apply_verdict_style(
    cell: Any,
    verdict: str,
) -> None:
    """
    Applies PASS, FAIL, or REVIEW fill style to a verdict cell.
    """

    normalized_verdict = verdict.upper()

    if normalized_verdict == "PASS":
        cell.fill = PASS_FILL

    elif normalized_verdict == "FAIL":
        cell.fill = FAIL_FILL

    else:
        cell.fill = REVIEW_FILL

    cell.font = Font(
        bold=True,
        color="1F1F1F",
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )