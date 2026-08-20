from collections import Counter
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

TITLE_FONT = Font(
    bold=True,
    color="FFFFFF",
    size=14,
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)


def generate_excel_report(
    output_path: Path,
    project_name: str,
    core: str,
    build_mode: str,
    project_root: Path,
    makefile_data: dict[str, Any],
    source_file_count: int,
    findings: Iterable[dict[str, Any]],
) -> Path:
    """
    Generates an Excel report with:

    - Summary
    - Compiler Switches
    - Preprocessor Conditions
    """

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    findings_list = list(findings)

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    switches_sheet = workbook.create_sheet(
        title="Compiler Switches"
    )

    conditions_sheet = workbook.create_sheet(
        title="Preprocessor Conditions"
    )

    filtered_switches_sheet = workbook.create_sheet(
    title="Filtered Compiler Switches"
    )

    excluded_conditions_sheet = workbook.create_sheet(
    title="Excluded Conditions"
    )

    relevant_switch_summary_sheet = workbook.create_sheet(
        title="Relevant Switch Summary"
    )

    _write_summary_sheet(
        worksheet=summary_sheet,
        project_name=project_name,
        core=core,
        build_mode=build_mode,
        project_root=project_root,
        source_file_count=source_file_count,
        findings=findings_list,
    )

    _write_compiler_switches_sheet(
        worksheet=switches_sheet,
        project_name=project_name,
        core=core,
        build_mode=build_mode,
        makefile_data=makefile_data,
        source_file_count=source_file_count,
    )

    _write_preprocessor_conditions_sheet(
        worksheet=conditions_sheet,
        project_name=project_name,
        core=core,
        build_mode=build_mode,
        findings=findings_list,
    )

    _write_filtered_compiler_switches_sheet(
        worksheet=filtered_switches_sheet,
        findings=findings_list,
    )

    _write_excluded_conditions_sheet(
        worksheet=excluded_conditions_sheet,
        findings=findings_list,
    )

    _write_relevant_switch_summary_sheet(
        worksheet=relevant_switch_summary_sheet,
        findings=findings_list,
    )

    workbook.save(output_path)
    workbook.close()

    return output_path


def create_compiler_switches_report(
    report_path: Path,
    findings: Iterable[dict[str, Any]],
) -> Path:
    """
    Simplified report-generation function used by main.py.

    This preserves compatibility with the current main.py integration.
    """

    return generate_excel_report(
        output_path=report_path,
        project_name="Release_DMS",
        core="Core1",
        build_mode="Release",
        project_root=Path("."),
        makefile_data={},
        source_file_count=0,
        findings=findings,
    )


def _write_summary_sheet(
    worksheet: Worksheet,
    project_name: str,
    core: str,
    build_mode: str,
    project_root: Path,
    source_file_count: int,
    findings: list[dict[str, Any]],
) -> None:
    """
    Writes the Summary worksheet with project metadata, global
    metrics, exclusion counts, and category distributions.
    """

    total_directives = len(findings)

    relevant_findings = [
        finding
        for finding in findings
        if finding.get("is_relevant", False)
    ]

    excluded_findings = [
        finding
        for finding in findings
        if finding.get("filter_reason", "")
    ]

    pending_findings = [
        finding
        for finding in findings
        if (
            finding.get("category", "OTHER") == "OTHER"
            and not finding.get("filter_reason", "")
        )
    ]

    exclusion_counts = _count_exclusions_by_reason(
        findings=findings,
    )

    category_counts = _count_findings_by_category(
        findings=findings,
    )

    relevant_category_counts = _count_findings_by_category(
        findings=relevant_findings,
    )

    worksheet["A1"] = "Metric"
    worksheet["B1"] = "Value"

    _format_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=2,
    )

    summary_rows = [
        ("Project", project_name),
        ("Core", core),
        ("Build mode", build_mode),
        ("Project root", str(project_root)),
        ("Source files analyzed", source_file_count),
        ("Total preprocessor directives analyzed", total_directives),
        ("Relevant compiler switches", len(relevant_findings)),
        ("Excluded conditions", len(excluded_findings)),
        (
            "Pending OTHER conditions for review",
            len(pending_findings),
        ),
    ]

    for row_number, (metric, value) in enumerate(
        summary_rows,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=metric,
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=value,
        )

    exclusions_start_row = 13

    _write_summary_section_header(
        worksheet=worksheet,
        row_number=exclusions_start_row,
        title="Exclusions by Reason",
    )

    worksheet.cell(
        row=exclusions_start_row + 1,
        column=1,
        value="Filter Reason",
    )

    worksheet.cell(
        row=exclusions_start_row + 1,
        column=2,
        value="Count",
    )

    _format_header_row(
        worksheet=worksheet,
        row_number=exclusions_start_row + 1,
        column_count=2,
    )

    exclusion_reasons = [
        "Header guard",
        "MemMap section marker",
        "Toolchain or architecture condition",
        "Vendor CMSIS configuration condition",
        "CMSIS framework condition",
        "Platform capability condition",
        "Static-analysis condition",
        "Test framework or instrumentation condition",
        "Generated or internal test condition",
        "Generated configuration-variant condition",
    ]

    exclusion_row = exclusions_start_row + 2

    for reason in exclusion_reasons:
        worksheet.cell(
            row=exclusion_row,
            column=1,
            value=reason,
        )

        worksheet.cell(
            row=exclusion_row,
            column=2,
            value=exclusion_counts.get(reason, 0),
        )

        exclusion_row += 1

    worksheet.cell(
        row=exclusion_row,
        column=1,
        value="Total excluded conditions",
    )

    worksheet.cell(
        row=exclusion_row,
        column=2,
        value=len(excluded_findings),
    )

    category_start_row = exclusion_row + 3

    _write_summary_section_header(
        worksheet=worksheet,
        row_number=category_start_row,
        title="Detected Conditions by Category",
    )

    worksheet.cell(
        row=category_start_row + 1,
        column=1,
        value="Category",
    )

    worksheet.cell(
        row=category_start_row + 1,
        column=2,
        value="Count",
    )

    _format_header_row(
        worksheet=worksheet,
        row_number=category_start_row + 1,
        column_count=2,
    )

    categories = [
        "DEBUG",
        "TEST",
        "INTEGRATION",
        "FEATURE",
        "OTHER",
    ]

    category_row = category_start_row + 2

    for category in categories:
        worksheet.cell(
            row=category_row,
            column=1,
            value=category,
        )

        worksheet.cell(
            row=category_row,
            column=2,
            value=category_counts.get(category, 0),
        )

        category_row += 1

    relevant_category_start_row = category_row + 2

    _write_summary_section_header(
        worksheet=worksheet,
        row_number=relevant_category_start_row,
        title="Relevant Compiler Switches by Category",
    )

    worksheet.cell(
        row=relevant_category_start_row + 1,
        column=1,
        value="Category",
    )

    worksheet.cell(
        row=relevant_category_start_row + 1,
        column=2,
        value="Relevant Count",
    )

    _format_header_row(
        worksheet=worksheet,
        row_number=relevant_category_start_row + 1,
        column_count=2,
    )

    relevant_category_row = relevant_category_start_row + 2

    for category in categories[:-1]:
        worksheet.cell(
            row=relevant_category_row,
            column=1,
            value=category,
        )

        worksheet.cell(
            row=relevant_category_row,
            column=2,
            value=relevant_category_counts.get(category, 0),
        )

        relevant_category_row += 1

    worksheet.cell(
        row=relevant_category_row,
        column=1,
        value="TOTAL",
    )

    worksheet.cell(
        row=relevant_category_row,
        column=2,
        value=len(relevant_findings),
    )

    worksheet.column_dimensions["A"].width = 50
    worksheet.column_dimensions["B"].width = 80
    worksheet.freeze_panes = "A2"


def _write_summary_section_header(
    worksheet: Worksheet,
    row_number: int,
    title: str,
) -> None:
    """
    Writes a formatted section title in the Summary worksheet.
    """

    worksheet.cell(
        row=row_number,
        column=1,
        value=title,
    )

    worksheet.merge_cells(
        start_row=row_number,
        start_column=1,
        end_row=row_number,
        end_column=2,
    )

    cell = worksheet.cell(
        row=row_number,
        column=1,
    )

    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )


def _count_exclusions_by_reason(
    findings: list[dict[str, Any]],
) -> Counter[str]:
    """
    Counts findings grouped by their non-empty filter reason.
    """

    return Counter(
        str(finding.get("filter_reason", ""))
        for finding in findings
        if finding.get("filter_reason", "")
    )


def _count_findings_by_category(
    findings: list[dict[str, Any]],
) -> Counter[str]:
    """
    Counts findings grouped by classifier category.
    """

    return Counter(
        str(finding.get("category", "OTHER"))
        for finding in findings
    )

def _write_compiler_switches_sheet(
    worksheet: Worksheet,
    project_name: str,
    core: str,
    build_mode: str,
    makefile_data: dict[str, Any],
    source_file_count: int,
) -> None:
    """
    Writes build/compiler data extracted from the Makefile.
    """

    headers = [
        "Project",
        "Core",
        "Build Mode",
        "Source Files",
        "Makefile Parameter",
        "Value",
        "Makefile Source",
        "Build Configuration",
        "Detected",
        "Entry Type",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

    _format_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=len(headers),
    )

    report_entries = list(
    makefile_data.items()
)[:5]

    for row_number, (parameter, value) in enumerate(
    report_entries,
    start=2,
):
        display_value = _format_value(value)

        worksheet.cell(row=row_number, column=1, value=project_name)
        worksheet.cell(row=row_number, column=2, value=core)
        worksheet.cell(row=row_number, column=3, value=build_mode)
        worksheet.cell(
            row=row_number,
            column=4,
            value=source_file_count,
        )
        worksheet.cell(
            row=row_number,
            column=5,
            value=str(parameter),
        )
        worksheet.cell(
            row=row_number,
            column=6,
            value=display_value,
        )
        worksheet.cell(
            row=row_number,
            column=7,
            value="Makefile",
        )
        worksheet.cell(
            row=row_number,
            column=8,
            value=build_mode,
        )
        worksheet.cell(
            row=row_number,
            column=9,
            value="Yes",
        )
        worksheet.cell(
            row=row_number,
            column=10,
            value=_get_entry_type(parameter),
        )

    _format_table(
        worksheet=worksheet,
        column_widths={
            "A": 24,
            "B": 14,
            "C": 16,
            "D": 14,
            "E": 35,
            "F": 65,
            "G": 18,
            "H": 22,
            "I": 14,
            "J": 22,
        },
    )


def _write_preprocessor_conditions_sheet(
    worksheet: Worksheet,
    project_name: str,
    core: str,
    build_mode: str,
    findings: list[dict[str, Any]],
) -> None:
    """
    Writes C/C++ preprocessor directives found by preprocessor_parser.py.
    """

    headers = [
    "Source File",
    "File Name",
    "Line",
    "Directive",
    "Expression",
    "Macros",
    ]

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

    _format_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=len(headers),
    )

    for row_number, finding in enumerate(
        findings,
        start=2,
    ):
        macros = finding.get("macros", [])

        if isinstance(macros, list):
            macros_text = ", ".join(
                str(macro)
                for macro in macros
            )
        else:
            macros_text = str(macros)

        # These names match src/preprocessor_parser.py.
        worksheet.cell(
            row=row_number,
            column=1,
            value=finding.get("path", ""),
        )
        worksheet.cell(
            row=row_number,
            column=2,
            value=finding.get("file_name", ""),
        )
        worksheet.cell(
            row=row_number,
            column=3,
            value=finding.get("line_number", ""),
        )
        worksheet.cell(
            row=row_number,
            column=4,
            value=finding.get("directive", ""),
        )
        worksheet.cell(
            row=row_number,
            column=5,
            value=finding.get("expression", ""),
        )
        worksheet.cell(
            row=row_number,
            column=6,
            value=macros_text,
        )

    _format_table(
        worksheet=worksheet,
        column_widths={
            "A": 80,
            "B": 30,
            "C": 12,
            "D": 16,
            "E": 85,
            "F": 65,
        },
    )


def _format_header_row(
    worksheet: Worksheet,
    row_number: int,
    column_count: int,
) -> None:
    """
    Applies common formatting to worksheet headers.
    """

    for column_number in range(1, column_count + 1):
        cell = worksheet.cell(
            row=row_number,
            column=column_number,
        )

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _format_table(
    worksheet: Worksheet,
    column_widths: dict[str, int],
) -> None:
    """
    Formats an Excel worksheet containing tabular data.
    """

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = (
            f"A1:{worksheet.cell(1, worksheet.max_column).coordinate}"
            f"{worksheet.max_row}"
        )

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def _format_value(value: Any) -> str:
    """
    Converts Makefile values into readable Excel text.
    """

    if isinstance(value, list):
        return " ".join(
            str(item)
            for item in value
        )

    if isinstance(value, tuple):
        return " ".join(
            str(item)
            for item in value
        )

    if isinstance(value, dict):
        return "; ".join(
            f"{key}={item}"
            for key, item in value.items()
        )

    return str(value)


def _get_entry_type(parameter: Any) -> str:
    """
    Classifies a Makefile entry for report readability.
    """

    parameter_name = str(parameter).lower()

    if "define" in parameter_name:
        return "Compiler definition"

    if "flag" in parameter_name:
        return "Compiler flag"

    if "optimization" in parameter_name:
        return "Optimization setting"

    if "compiler" in parameter_name:
        return "Compiler setting"

    return "Build setting"

def _write_filtered_compiler_switches_sheet(
    worksheet: Worksheet,
    findings: list[dict[str, Any]],
) -> None:
    """
    Writes only relevant functional compiler switches.

    Relevant findings are identified by switch_classifier.py.
    """

    headers = [
        "Source File",
        "File Name",
        "Line",
        "Directive",
        "Expression",
        "Macros",
        "Category",
        "Matched Keywords",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    relevant_findings = [
        finding
        for finding in findings
        if finding.get("is_relevant", False)
    ]

    for row_number, finding in enumerate(
        relevant_findings,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=str(finding.get("path", "")),
        )
        worksheet.cell(
            row=row_number,
            column=2,
            value=str(finding.get("file_name", "")),
        )
        worksheet.cell(
            row=row_number,
            column=3,
            value=finding.get("line_number", ""),
        )
        worksheet.cell(
            row=row_number,
            column=4,
            value=str(finding.get("directive", "")),
        )
        worksheet.cell(
            row=row_number,
            column=5,
            value=str(finding.get("expression", "")),
        )
        worksheet.cell(
            row=row_number,
            column=6,
            value=_format_list_value(
                finding.get("macros", [])
            ),
        )
        worksheet.cell(
            row=row_number,
            column=7,
            value=str(finding.get("category", "OTHER")),
        )
        worksheet.cell(
            row=row_number,
            column=8,
            value=_format_list_value(
                finding.get("matched_keywords", [])
            ),
        )

    _format_table(
        worksheet=worksheet,
        column_widths={
            "A": 80,
            "B": 30,
            "C": 12,
            "D": 16,
            "E": 85,
            "F": 65,
            "G": 18,
            "H": 30,
        },
    )


def _write_excluded_conditions_sheet(
    worksheet: Worksheet,
    findings: list[dict[str, Any]],
) -> None:
    """
    Writes non-relevant conditions and the reason why they were
    excluded from the functional-switch report.
    """

    headers = [
        "Source File",
        "File Name",
        "Line",
        "Directive",
        "Expression",
        "Macros",
        "Category",
        "Filter Reason",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    excluded_findings = [
        finding
        for finding in findings
        if not finding.get("is_relevant", False)
    ]

    for row_number, finding in enumerate(
        excluded_findings,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=str(finding.get("path", "")),
        )
        worksheet.cell(
            row=row_number,
            column=2,
            value=str(finding.get("file_name", "")),
        )
        worksheet.cell(
            row=row_number,
            column=3,
            value=finding.get("line_number", ""),
        )
        worksheet.cell(
            row=row_number,
            column=4,
            value=str(finding.get("directive", "")),
        )
        worksheet.cell(
            row=row_number,
            column=5,
            value=str(finding.get("expression", "")),
        )
        worksheet.cell(
            row=row_number,
            column=6,
            value=_format_list_value(
                finding.get("macros", [])
            ),
        )
        worksheet.cell(
            row=row_number,
            column=7,
            value=str(finding.get("category", "OTHER")),
        )
        worksheet.cell(
            row=row_number,
            column=8,
            value=str(finding.get("filter_reason", "")),
        )

    _format_table(
        worksheet=worksheet,
        column_widths={
            "A": 80,
            "B": 30,
            "C": 12,
            "D": 16,
            "E": 85,
            "F": 65,
            "G": 18,
            "H": 38,
        },
    )


def _write_headers(
    worksheet: Worksheet,
    headers: list[str],
) -> None:
    """
    Writes and formats a header row beginning in row 1.
    """

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

    _format_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=len(headers),
    )


def _format_list_value(value: Any) -> str:
    """
    Converts list-like values into comma-separated Excel text.
    """

    if isinstance(value, list):
        return ", ".join(
            str(item)
            for item in value
        )

    if isinstance(value, tuple):
        return ", ".join(
            str(item)
            for item in value
        )

def _write_relevant_switch_summary_sheet(
    worksheet: Worksheet,
    findings: list[dict[str, Any]],
) -> None:
    """
    Writes one row per unique relevant compiler switch.

    Detailed occurrences remain available in the
    Filtered Compiler Switches worksheet.
    """

    headers = [
        "Category",
        "Primary Macro",
        "Occurrences",
        "Files Affected",
        "Example Expression",
        "Example Source File",
        "Example Line",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    grouped_switches = _group_relevant_switches(
        findings=findings,
    )

    for row_number, switch_data in enumerate(
        grouped_switches,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=switch_data["category"],
        )
        worksheet.cell(
            row=row_number,
            column=2,
            value=switch_data["primary_macro"],
        )
        worksheet.cell(
            row=row_number,
            column=3,
            value=switch_data["occurrences"],
        )
        worksheet.cell(
            row=row_number,
            column=4,
            value=switch_data["files_affected"],
        )
        worksheet.cell(
            row=row_number,
            column=5,
            value=switch_data["example_expression"],
        )
        worksheet.cell(
            row=row_number,
            column=6,
            value=switch_data["example_source_file"],
        )
        worksheet.cell(
            row=row_number,
            column=7,
            value=switch_data["example_line"],
        )

    _format_table(
        worksheet=worksheet,
        column_widths={
            "A": 18,
            "B": 55,
            "C": 14,
            "D": 16,
            "E": 85,
            "F": 80,
            "G": 14,
        },
    )


def _group_relevant_switches(
    findings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Groups relevant findings by category and primary macro.

    The first parser macro is used as the primary switch macro.
    This is normally the macro controlling the preprocessor
    expression, for example BSWM_ENABLE_CANSM or FEATURE_X.
    """

    grouped_data: dict[
        tuple[str, str],
        dict[str, Any],
    ] = {}

    for finding in findings:
        if not finding.get("is_relevant", False):
            continue

        category = str(
            finding.get("category", "OTHER")
        )

        primary_macro = _get_primary_macro(
            finding=finding,
        )

        group_key = (
            category,
            primary_macro,
        )

        if group_key not in grouped_data:
            grouped_data[group_key] = {
                "category": category,
                "primary_macro": primary_macro,
                "occurrences": 0,
                "source_files": set(),
                "example_expression": str(
                    finding.get("expression", "")
                ),
                "example_source_file": str(
                    finding.get("path", "")
                ),
                "example_line": finding.get(
                    "line_number",
                    "",
                ),
            }

        grouped_data[group_key]["occurrences"] += 1

        source_file = str(
            finding.get("path", "")
        )

        if source_file:
            grouped_data[group_key][
                "source_files"
            ].add(source_file)

    grouped_switches = []

    for switch_data in grouped_data.values():
        grouped_switches.append(
            {
                "category": switch_data["category"],
                "primary_macro": switch_data[
                    "primary_macro"
                ],
                "occurrences": switch_data[
                    "occurrences"
                ],
                "files_affected": len(
                    switch_data["source_files"]
                ),
                "example_expression": switch_data[
                    "example_expression"
                ],
                "example_source_file": switch_data[
                    "example_source_file"
                ],
                "example_line": switch_data[
                    "example_line"
                ],
            }
        )

    return sorted(
        grouped_switches,
        key=lambda item: (
            -int(item["occurrences"]),
            str(item["category"]),
            str(item["primary_macro"]),
        ),
    )


def _get_primary_macro(
    finding: dict[str, Any],
) -> str:
    """
    Returns the first parser macro as the representative switch.

    Falls back to the full expression if no macro was extracted.
    """

    macros = finding.get("macros", [])

    if isinstance(macros, list) and macros:
        return str(macros[0])

    if isinstance(macros, tuple) and macros:
        return str(macros[0])

    expression = str(
        finding.get("expression", "")
    )

    return expression or "UNRESOLVED_EXPRESSION"

    return str(value)